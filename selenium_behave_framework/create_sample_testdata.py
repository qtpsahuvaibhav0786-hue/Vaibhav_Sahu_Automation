#!/usr/bin/env python3
"""
Script to generate sample test data Excel file for keyword-driven testing
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path


def create_sample_test_data():
    """Create sample test data Excel file"""

    # Create workbook
    wb = openpyxl.Workbook()

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def style_header(sheet, row=1):
        """Apply header styling"""
        for cell in sheet[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    # ==================== Master Sheet ====================
    master = wb.active
    master.title = "Master"
    master.append(["TestCaseID", "Execute", "ScreenFlow", "Description"])
    style_header(master)

    # Sample test cases
    master.append(["TC001", "Yes", "Login,Inventory", "Login and view products"])
    master.append(["TC002", "Yes", "Login,AddToCart,Checkout", "Complete purchase flow"])
    master.append(["TC003", "Yes", "Login,InvalidLogin", "Test invalid login"])
    master.append(["TC004", "No", "Login", "Skip this test"])

    # Adjust column widths
    master.column_dimensions['A'].width = 15
    master.column_dimensions['B'].width = 10
    master.column_dimensions['C'].width = 30
    master.column_dimensions['D'].width = 40

    # ==================== Login Screen ====================
    login = wb.create_sheet("Login")
    login.append(["StepNo", "Keyword", "Locator", "LocatorValue", "TestData", "Description"])
    style_header(login)

    login.append([1, "NAVIGATE", "", "", "https://www.saucedemo.com", "Open application URL"])
    login.append([2, "WAIT_FOR_ELEMENT", "id", "user-name", "", "Wait for login form"])
    login.append([3, "ENTER", "id", "user-name", "standard_user", "Enter username"])
    login.append([4, "ENTER", "id", "password", "secret_sauce", "Enter password"])
    login.append([5, "CLICK", "id", "login-button", "", "Click login button"])
    login.append([6, "WAIT_FOR_ELEMENT", "class", "inventory_container", "", "Wait for inventory page"])
    login.append([7, "VERIFY_URL", "", "", "inventory.html", "Verify on inventory page"])

    # Adjust column widths
    for col, width in [('A', 10), ('B', 20), ('C', 10), ('D', 30), ('E', 30), ('F', 35)]:
        login.column_dimensions[col].width = width

    # ==================== Invalid Login Screen ====================
    invalid_login = wb.create_sheet("InvalidLogin")
    invalid_login.append(["StepNo", "Keyword", "Locator", "LocatorValue", "TestData", "Description"])
    style_header(invalid_login)

    invalid_login.append([1, "NAVIGATE", "", "", "https://www.saucedemo.com", "Open application URL"])
    invalid_login.append([2, "ENTER", "id", "user-name", "invalid_user", "Enter invalid username"])
    invalid_login.append([3, "ENTER", "id", "password", "wrong_password", "Enter wrong password"])
    invalid_login.append([4, "CLICK", "id", "login-button", "", "Click login button"])
    invalid_login.append([5, "WAIT_FOR_VISIBLE", "css", "h3[data-test='error']", "", "Wait for error message"])
    invalid_login.append([6, "VERIFY_TEXT", "css", "h3[data-test='error']", "do not match", "Verify error message"])
    invalid_login.append([7, "SCREENSHOT", "", "", "invalid_login_error", "Capture error screenshot"])

    for col, width in [('A', 10), ('B', 20), ('C', 10), ('D', 30), ('E', 30), ('F', 35)]:
        invalid_login.column_dimensions[col].width = width

    # ==================== Inventory Screen ====================
    inventory = wb.create_sheet("Inventory")
    inventory.append(["StepNo", "Keyword", "Locator", "LocatorValue", "TestData", "Description"])
    style_header(inventory)

    inventory.append([1, "VERIFY_ELEMENT_VISIBLE", "class", "inventory_list", "", "Verify products displayed"])
    inventory.append([2, "VERIFY_TITLE", "", "", "Swag Labs", "Verify page title"])
    inventory.append([3, "SCREENSHOT", "", "", "inventory_page", "Capture inventory page"])

    for col, width in [('A', 10), ('B', 25), ('C', 10), ('D', 30), ('E', 25), ('F', 35)]:
        inventory.column_dimensions[col].width = width

    # ==================== AddToCart Screen ====================
    add_to_cart = wb.create_sheet("AddToCart")
    add_to_cart.append(["StepNo", "Keyword", "Locator", "LocatorValue", "TestData", "Description"])
    style_header(add_to_cart)

    add_to_cart.append([1, "CLICK", "id", "add-to-cart-sauce-labs-backpack", "", "Add Backpack to cart"])
    add_to_cart.append([2, "CLICK", "id", "add-to-cart-sauce-labs-bike-light", "", "Add Bike Light to cart"])
    add_to_cart.append([3, "VERIFY_TEXT", "class", "shopping_cart_badge", "2", "Verify cart count"])
    add_to_cart.append([4, "CLICK", "class", "shopping_cart_link", "", "Click cart icon"])
    add_to_cart.append([5, "WAIT_FOR_ELEMENT", "class", "cart_list", "", "Wait for cart page"])
    add_to_cart.append([6, "VERIFY_ELEMENT_VISIBLE", "xpath", "//div[text()='Sauce Labs Backpack']", "", "Verify Backpack in cart"])

    for col, width in [('A', 10), ('B', 25), ('C', 10), ('D', 40), ('E', 20), ('F', 35)]:
        add_to_cart.column_dimensions[col].width = width

    # ==================== Checkout Screen ====================
    checkout = wb.create_sheet("Checkout")
    checkout.append(["StepNo", "Keyword", "Locator", "LocatorValue", "TestData", "Description"])
    style_header(checkout)

    checkout.append([1, "CLICK", "id", "checkout", "", "Click Checkout button"])
    checkout.append([2, "WAIT_FOR_ELEMENT", "id", "first-name", "", "Wait for checkout form"])
    checkout.append([3, "ENTER", "id", "first-name", "John", "Enter first name"])
    checkout.append([4, "ENTER", "id", "last-name", "Doe", "Enter last name"])
    checkout.append([5, "ENTER", "id", "postal-code", "12345", "Enter postal code"])
    checkout.append([6, "CLICK", "id", "continue", "", "Click Continue"])
    checkout.append([7, "WAIT_FOR_ELEMENT", "id", "finish", "", "Wait for finish button"])
    checkout.append([8, "VERIFY_ELEMENT_VISIBLE", "class", "summary_total_label", "", "Verify total displayed"])
    checkout.append([9, "CLICK", "id", "finish", "", "Click Finish"])
    checkout.append([10, "WAIT_FOR_ELEMENT", "class", "complete-header", "", "Wait for confirmation"])
    checkout.append([11, "VERIFY_TEXT", "class", "complete-header", "Thank you", "Verify order success"])
    checkout.append([12, "SCREENSHOT", "", "", "order_complete", "Capture confirmation"])

    for col, width in [('A', 10), ('B', 25), ('C', 10), ('D', 35), ('E', 20), ('F', 35)]:
        checkout.column_dimensions[col].width = width

    # ==================== Keywords Reference Sheet ====================
    keywords_ref = wb.create_sheet("Keywords_Reference")
    keywords_ref.append(["Category", "Keyword", "Description", "Parameters"])
    style_header(keywords_ref)

    keywords_data = [
        ["Navigation", "NAVIGATE", "Navigate to URL", "TestData: URL"],
        ["Navigation", "REFRESH", "Refresh current page", "None"],
        ["Navigation", "GO_BACK", "Navigate back", "None"],
        ["Navigation", "GO_FORWARD", "Navigate forward", "None"],
        ["Input", "ENTER", "Enter text into element", "Locator + TestData"],
        ["Input", "CLEAR", "Clear input field", "Locator"],
        ["Input", "CLEAR_AND_ENTER", "Clear and enter text", "Locator + TestData"],
        ["Click", "CLICK", "Click element", "Locator"],
        ["Click", "DOUBLE_CLICK", "Double-click element", "Locator"],
        ["Click", "RIGHT_CLICK", "Right-click element", "Locator"],
        ["Click", "JS_CLICK", "Click using JavaScript", "Locator"],
        ["Dropdown", "SELECT", "Select by visible text", "Locator + TestData"],
        ["Dropdown", "SELECT_BY_VALUE", "Select by value attribute", "Locator + TestData"],
        ["Dropdown", "SELECT_BY_INDEX", "Select by index", "Locator + TestData (index)"],
        ["Checkbox", "CHECK", "Check checkbox", "Locator"],
        ["Checkbox", "UNCHECK", "Uncheck checkbox", "Locator"],
        ["Verify", "VERIFY_TEXT", "Verify element text", "Locator + TestData"],
        ["Verify", "VERIFY_TITLE", "Verify page title", "TestData"],
        ["Verify", "VERIFY_URL", "Verify current URL", "TestData"],
        ["Verify", "VERIFY_ELEMENT_VISIBLE", "Verify element is visible", "Locator"],
        ["Verify", "VERIFY_ELEMENT_PRESENT", "Verify element in DOM", "Locator"],
        ["Wait", "WAIT", "Wait for seconds", "TestData (seconds)"],
        ["Wait", "WAIT_FOR_ELEMENT", "Wait for element present", "Locator"],
        ["Wait", "WAIT_FOR_VISIBLE", "Wait for element visible", "Locator"],
        ["Wait", "WAIT_FOR_CLICKABLE", "Wait for element clickable", "Locator"],
        ["Keyboard", "PRESS_KEY", "Press keyboard key", "Locator + TestData (key)"],
        ["Keyboard", "PRESS_ENTER", "Press Enter key", "Locator (optional)"],
        ["Mouse", "HOVER", "Hover over element", "Locator"],
        ["Mouse", "SCROLL_TO", "Scroll to element", "Locator"],
        ["Frame", "SWITCH_TO_FRAME", "Switch to iframe", "TestData (name/id/index)"],
        ["Frame", "SWITCH_TO_DEFAULT", "Switch to default content", "None"],
        ["Alert", "ACCEPT_ALERT", "Accept alert dialog", "None"],
        ["Alert", "DISMISS_ALERT", "Dismiss alert dialog", "None"],
        ["Data", "GET_TEXT", "Get element text", "Locator"],
        ["Data", "STORE_TEXT", "Store text in variable", "Locator + TestData (var name)"],
        ["Screenshot", "SCREENSHOT", "Take screenshot", "TestData (filename)"],
        ["Utility", "LOG", "Log message", "TestData (message)"],
        ["Utility", "PAUSE", "Pause execution", "TestData (seconds)"],
    ]

    for row in keywords_data:
        keywords_ref.append(row)

    for col, width in [('A', 15), ('B', 25), ('C', 35), ('D', 30)]:
        keywords_ref.column_dimensions[col].width = width

    # Save workbook
    test_data_dir = Path(__file__).parent / "test_data"
    test_data_dir.mkdir(exist_ok=True)
    output_path = test_data_dir / "test_data.xlsx"
    wb.save(output_path)

    print(f"Sample test data created: {output_path}")
    print("\nSheets created:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")

    return output_path


if __name__ == "__main__":
    create_sample_test_data()
