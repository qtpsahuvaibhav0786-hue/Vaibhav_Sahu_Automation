"""
Excel Reader utility for reading test data from Excel files
Supports keyword-driven and data-driven testing
"""

from typing import Dict, List, Any, Optional
from pathlib import Path
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..config.config import Config


class ExcelReader:
    """
    Excel Reader class for reading test data from Excel files
    Supports multiple sheets and various data formats
    """

    def __init__(self, file_path: Path = None):
        """
        Initialize Excel Reader

        Args:
            file_path: Path to Excel file
        """
        self.file_path = file_path or Config.get_test_data_path()
        self._workbook: Optional[Workbook] = None

    def _load_workbook(self) -> Workbook:
        """Load workbook if not already loaded"""
        if self._workbook is None:
            if not self.file_path.exists():
                raise FileNotFoundError(f"Test data file not found: {self.file_path}")
            self._workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        return self._workbook

    def get_sheet_names(self) -> List[str]:
        """Get all sheet names in workbook"""
        workbook = self._load_workbook()
        return workbook.sheetnames

    def get_sheet(self, sheet_name: str) -> Worksheet:
        """Get specific sheet by name"""
        workbook = self._load_workbook()
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        return workbook[sheet_name]

    def read_all_data(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Read all data from a sheet as list of dictionaries
        First row is treated as headers

        Args:
            sheet_name: Name of the sheet to read

        Returns:
            List of dictionaries with column headers as keys
        """
        sheet = self.get_sheet(sheet_name)
        data = []

        # Get headers from first row
        headers = [cell.value for cell in sheet[1] if cell.value is not None]

        # Read data rows
        for row_idx in range(2, sheet.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                row_data[header] = cell_value
            # Only add row if it has at least one non-None value
            if any(v is not None for v in row_data.values()):
                data.append(row_data)

        return data

    def read_test_cases(self, sheet_name: str = "TestCases") -> List[Dict[str, Any]]:
        """
        Read test cases from the TestCases sheet

        Args:
            sheet_name: Name of test cases sheet

        Returns:
            List of test case dictionaries
        """
        return self.read_all_data(sheet_name)

    def read_keywords(self, sheet_name: str = "Keywords") -> List[Dict[str, Any]]:
        """
        Read keyword steps from a sheet
        Expected columns: StepNo, Keyword, Locator, LocatorValue, TestData, Description

        Args:
            sheet_name: Name of keywords sheet

        Returns:
            List of keyword step dictionaries
        """
        data = self.read_all_data(sheet_name)
        # Convert to standard format
        keywords = []
        for row in data:
            keyword = {
                'step_no': row.get('StepNo') or row.get('Step'),
                'keyword': row.get('Keyword') or row.get('Action'),
                'locator_type': row.get('Locator') or row.get('LocatorType'),
                'locator_value': row.get('LocatorValue') or row.get('Element'),
                'test_data': row.get('TestData') or row.get('Value') or row.get('Data'),
                'description': row.get('Description') or row.get('Comment') or ''
            }
            if keyword['keyword']:  # Only add if keyword is present
                keywords.append(keyword)
        return keywords

    def read_test_data(self, sheet_name: str, test_case_id: str) -> Dict[str, Any]:
        """
        Read test data for a specific test case

        Args:
            sheet_name: Name of test data sheet
            test_case_id: Test case identifier

        Returns:
            Dictionary of test data
        """
        data = self.read_all_data(sheet_name)
        for row in data:
            if row.get('TestCaseID') == test_case_id or row.get('TC_ID') == test_case_id:
                return row
        return {}

    def read_cell(self, sheet_name: str, row: int, column: int) -> Any:
        """
        Read a specific cell value

        Args:
            sheet_name: Name of sheet
            row: Row number (1-indexed)
            column: Column number (1-indexed)

        Returns:
            Cell value
        """
        sheet = self.get_sheet(sheet_name)
        return sheet.cell(row=row, column=column).value

    def read_column(self, sheet_name: str, column: int) -> List[Any]:
        """
        Read all values in a column

        Args:
            sheet_name: Name of sheet
            column: Column number (1-indexed)

        Returns:
            List of column values
        """
        sheet = self.get_sheet(sheet_name)
        return [sheet.cell(row=i, column=column).value for i in range(1, sheet.max_row + 1)]

    def read_row(self, sheet_name: str, row: int) -> List[Any]:
        """
        Read all values in a row

        Args:
            sheet_name: Name of sheet
            row: Row number (1-indexed)

        Returns:
            List of row values
        """
        sheet = self.get_sheet(sheet_name)
        return [sheet.cell(row=row, column=i).value for i in range(1, sheet.max_column + 1)]

    def get_row_count(self, sheet_name: str) -> int:
        """Get total number of rows in sheet"""
        sheet = self.get_sheet(sheet_name)
        return sheet.max_row

    def get_column_count(self, sheet_name: str) -> int:
        """Get total number of columns in sheet"""
        sheet = self.get_sheet(sheet_name)
        return sheet.max_column

    def get_test_scenarios(self, sheet_name: str = "Master") -> List[Dict[str, Any]]:
        """
        Read test scenarios from master sheet
        Expected format: TestCaseID, Execute, ScreenFlow, Description

        Args:
            sheet_name: Name of master sheet

        Returns:
            List of test scenarios to execute
        """
        data = self.read_all_data(sheet_name)
        scenarios = []
        for row in data:
            execute = str(row.get('Execute', 'No')).lower()
            if execute in ('yes', 'y', 'true', '1'):
                scenarios.append({
                    'test_case_id': row.get('TestCaseID') or row.get('TC_ID'),
                    'screen_flow': row.get('ScreenFlow') or row.get('Screens'),
                    'description': row.get('Description') or ''
                })
        return scenarios

    def close(self):
        """Close the workbook"""
        if self._workbook:
            self._workbook.close()
            self._workbook = None

    def __enter__(self):
        """Context manager entry"""
        self._load_workbook()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit"""
        self.close()
