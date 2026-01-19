"""
Script to create sample test data Excel file
This creates a template Excel file with:
1. Master Sheet - for test case selection and screen flow definition
2. Sample screen sheets - LoginScreen, DashboardScreen
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

def create_sample_test_data():
    """Create sample test data Excel file"""

    # Create new workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    locator_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    keyword_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    value_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    # ==================== MASTER SHEET ====================
    master_sheet = wb.create_sheet("MasterSheet")

    # Master sheet headers
    headers = ["TestCaseID", "Execute", "ScreenFlow", "Description"]
    for col, header in enumerate(headers, start=1):
        cell = master_sheet.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Master sheet test data
    master_data = [
        ["TC001", "Yes", "LoginScreen", "Valid login test case"],
        ["TC002", "No", "LoginScreen", "Invalid login test case"],
        ["TC003", "Yes", "LoginScreen,DashboardScreen", "Login and navigate to dashboard"],
        ["TC004", "Yes", "DashboardScreen", "Dashboard verification test"]
    ]

    for row_idx, data in enumerate(master_data, start=2):
        for col_idx, value in enumerate(data, start=1):
            cell = master_sheet.cell(row_idx, col_idx, value)
            cell.alignment = center_align

    # Set column widths
    master_sheet.column_dimensions['A'].width = 15
    master_sheet.column_dimensions['B'].width = 10
    master_sheet.column_dimensions['C'].width = 35
    master_sheet.column_dimensions['D'].width = 40

    # ==================== LOGIN SCREEN SHEET ====================
    login_sheet = wb.create_sheet("LoginScreen")

    # Row 1: Field names
    fields = ["URL", "Username", "Password", "LoginButton", "ErrorMessage"]
    for col, field in enumerate(fields, start=1):
        cell = login_sheet.cell(1, col, field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Row 2: XPath/Locators
    locators = [
        "https://demo.playwright.dev/todomvc",  # Example URL
        "//input[@id='username']",
        "//input[@id='password']",
        "//button[@id='login']",
        "//div[@class='error-message']"
    ]
    for col, locator in enumerate(locators, start=1):
        cell = login_sheet.cell(2, col, locator)
        cell.fill = locator_fill
        cell.alignment = center_align

    # Row 3-4: Test Case 1 - Keywords and Values
    keywords_tc1 = ["NAVIGATE", "ENTER", "ENTER", "CLICK", "VERIFY_TEXT"]
    values_tc1 = ["", "admin@test.com", "Admin@123", "", ""]

    for col, keyword in enumerate(keywords_tc1, start=1):
        cell = login_sheet.cell(3, col, keyword)
        cell.fill = keyword_fill
        cell.alignment = center_align

    for col, value in enumerate(values_tc1, start=1):
        cell = login_sheet.cell(4, col, value)
        cell.fill = value_fill
        cell.alignment = center_align

    # Row 5-6: Test Case 2 - Keywords and Values (Invalid login)
    keywords_tc2 = ["NAVIGATE", "ENTER", "ENTER", "CLICK", "VERIFY_TEXT"]
    values_tc2 = ["", "invalid@test.com", "WrongPass", "", "Invalid credentials"]

    for col, keyword in enumerate(keywords_tc2, start=1):
        cell = login_sheet.cell(5, col, keyword)
        cell.fill = keyword_fill
        cell.alignment = center_align

    for col, value in enumerate(values_tc2, start=1):
        cell = login_sheet.cell(6, col, value)
        cell.fill = value_fill
        cell.alignment = center_align

    # Set column widths
    for col in range(1, len(fields) + 1):
        login_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    # ==================== DASHBOARD SCREEN SHEET ====================
    dashboard_sheet = wb.create_sheet("DashboardScreen")

    # Row 1: Field names
    fields = ["WelcomeMessage", "ProfileIcon", "LogoutButton", "Title"]
    for col, field in enumerate(fields, start=1):
        cell = dashboard_sheet.cell(1, col, field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Row 2: XPath/Locators
    locators = [
        "//h1[@class='welcome']",
        "//div[@id='profile-icon']",
        "//button[@id='logout']",
        "Dashboard"
    ]
    for col, locator in enumerate(locators, start=1):
        cell = dashboard_sheet.cell(2, col, locator)
        cell.fill = locator_fill
        cell.alignment = center_align

    # Row 3-4: Test Case 1 - Keywords and Values
    keywords_tc1 = ["VERIFY_TEXT", "CLICK", "", "VERIFY_TITLE"]
    values_tc1 = ["Welcome", "", "", "Dashboard"]

    for col, keyword in enumerate(keywords_tc1, start=1):
        cell = dashboard_sheet.cell(3, col, keyword)
        cell.fill = keyword_fill
        cell.alignment = center_align

    for col, value in enumerate(values_tc1, start=1):
        cell = dashboard_sheet.cell(4, col, value)
        cell.fill = value_fill
        cell.alignment = center_align

    # Row 5-6: Test Case 2 - Keywords and Values (Logout flow)
    keywords_tc2 = ["WAIT_FOR_ELEMENT", "HOVER", "CLICK", ""]
    values_tc2 = ["3", "", "", ""]

    for col, keyword in enumerate(keywords_tc2, start=1):
        cell = dashboard_sheet.cell(5, col, keyword)
        cell.fill = keyword_fill
        cell.alignment = center_align

    for col, value in enumerate(values_tc2, start=1):
        cell = dashboard_sheet.cell(6, col, value)
        cell.fill = value_fill
        cell.alignment = center_align

    # Set column widths
    for col in range(1, len(fields) + 1):
        dashboard_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    # Save workbook
    test_data_dir = Path(__file__).parent / "test_data"
    test_data_dir.mkdir(exist_ok=True)
    file_path = test_data_dir / "TestData.xlsx"
    wb.save(file_path)

    print(f"✓ Sample test data file created successfully: {file_path}")
    print(f"\nFile Structure:")
    print(f"  - MasterSheet: Test case selection and screen flow")
    print(f"  - LoginScreen: Login page test data")
    print(f"  - DashboardScreen: Dashboard page test data")
    print(f"\nYou can now customize this file with your actual:")
    print(f"  - URLs and locators")
    print(f"  - Test cases and screen flows")
    print(f"  - Keywords and test data")

if __name__ == "__main__":
    create_sample_test_data()
