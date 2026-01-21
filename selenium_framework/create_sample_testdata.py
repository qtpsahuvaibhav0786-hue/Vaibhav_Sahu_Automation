"""
Sample Test Data Generator.
Creates a sample Excel file with test cases for the Selenium framework.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path


def create_sample_excel():
    """Create sample Excel file with test data."""

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create Master Sheet
    create_master_sheet(wb)

    # Create LoginScreen sheet
    create_login_screen(wb)

    # Create DashboardScreen sheet
    create_dashboard_screen(wb)

    # Save workbook
    excel_path = Path(__file__).parent / "test_data" / "TestData.xlsx"
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(excel_path)
    print(f"✓ Sample Excel file created: {excel_path}")
    print(f"\nThe Excel file contains:")
    print(f"  - Master sheet with test case execution control")
    print(f"  - LoginScreen with sample login test cases")
    print(f"  - DashboardScreen with sample dashboard tests")
    print(f"\nYou can now run tests using: python run_tests.py")


def create_master_sheet(wb):
    """Create Master sheet for test case management."""

    ws = wb.create_sheet("Master", 0)

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header row
    headers = ["TestCaseID", "Execute", "ScreenFlow", "Description"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Test case data
    test_cases = [
        ["TC001", "Yes", "LoginScreen", "Valid login test with standard user"],
        ["TC002", "Yes", "LoginScreen", "Invalid login test with wrong credentials"],
        ["TC003", "Yes", "LoginScreen,DashboardScreen", "Login and verify dashboard elements"],
        ["TC004", "No", "LoginScreen", "Login with locked out user (disabled)"],
    ]

    for row_num, test_case in enumerate(test_cases, 2):
        for col_num, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 50


def create_login_screen(wb):
    """Create LoginScreen sheet with login test data."""

    ws = wb.create_sheet("LoginScreen")

    # Define styles
    field_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    locator_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    keyword_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    value_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    header_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1: Field Names
    field_names = ["URL", "Username", "Password", "LoginButton", "ErrorMessage"]
    for col_num, field in enumerate(field_names, 1):
        cell = ws.cell(row=1, column=col_num, value=field)
        cell.fill = field_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Row 2: Locators
    locators = [
        "https://www.saucedemo.com",
        "id=user-name",
        "id=password",
        "id=login-button",
        "//h3[@data-test='error']"
    ]
    for col_num, locator in enumerate(locators, 1):
        cell = ws.cell(row=2, column=col_num, value=locator)
        cell.fill = locator_fill
        cell.font = Font(size=10)
        cell.alignment = Alignment(vertical="center")
        cell.border = border

    # Test Case 1: Valid Login (TC001)
    # Row 3: Keywords
    keywords_tc1 = ["NAVIGATE", "ENTER", "ENTER", "CLICK", ""]
    for col_num, keyword in enumerate(keywords_tc1, 1):
        cell = ws.cell(row=3, column=col_num, value=keyword)
        cell.fill = keyword_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Row 4: Values
    values_tc1 = ["", "standard_user", "secret_sauce", "", ""]
    for col_num, value in enumerate(values_tc1, 1):
        cell = ws.cell(row=4, column=col_num, value=value)
        cell.fill = value_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = border

    # Test Case 2: Invalid Login (TC002)
    # Row 5: Keywords
    keywords_tc2 = ["NAVIGATE", "ENTER", "ENTER", "CLICK", "VERIFY_TEXT"]
    for col_num, keyword in enumerate(keywords_tc2, 1):
        cell = ws.cell(row=5, column=col_num, value=keyword)
        cell.fill = keyword_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Row 6: Values
    values_tc2 = ["", "invalid_user", "wrong_password", "", "Epic sadface"]
    for col_num, value in enumerate(values_tc2, 1):
        cell = ws.cell(row=6, column=col_num, value=value)
        cell.fill = value_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = border

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30


def create_dashboard_screen(wb):
    """Create DashboardScreen sheet with dashboard verification data."""

    ws = wb.create_sheet("DashboardScreen")

    # Define styles
    field_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    locator_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    keyword_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    value_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    header_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1: Field Names
    field_names = ["PageTitle", "ProductsHeader", "CartIcon", "MenuButton"]
    for col_num, field in enumerate(field_names, 1):
        cell = ws.cell(row=1, column=col_num, value=field)
        cell.fill = field_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Row 2: Locators
    locators = [
        "",
        "//span[@class='title']",
        "//a[@class='shopping_cart_link']",
        "id=react-burger-menu-btn"
    ]
    for col_num, locator in enumerate(locators, 1):
        cell = ws.cell(row=2, column=col_num, value=locator)
        cell.fill = locator_fill
        cell.font = Font(size=10)
        cell.alignment = Alignment(vertical="center")
        cell.border = border

    # Test Case 3: Dashboard Verification (TC003)
    # Row 3: Keywords
    keywords_tc3 = ["VERIFY_TITLE", "VERIFY_TEXT", "VERIFY_ELEMENT", "VERIFY_ELEMENT"]
    for col_num, keyword in enumerate(keywords_tc3, 1):
        cell = ws.cell(row=3, column=col_num, value=keyword)
        cell.fill = keyword_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Row 4: Values
    values_tc3 = ["Swag Labs", "Products", "", ""]
    for col_num, value in enumerate(values_tc3, 1):
        cell = ws.cell(row=4, column=col_num, value=value)
        cell.fill = value_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = border

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25


if __name__ == "__main__":
    create_sample_excel()
