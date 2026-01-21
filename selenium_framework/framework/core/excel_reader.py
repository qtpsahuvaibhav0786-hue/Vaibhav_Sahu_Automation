"""
Excel Reader Module.
Handles reading test data from Excel files in the specified format.
"""

import openpyxl
from pathlib import Path
from framework.config.config import Config
from framework.utils.logger import Logger


class ExcelReader:
    """Reads test data from Excel files."""

    def __init__(self, excel_path=None):
        """
        Initialize Excel Reader.

        Args:
            excel_path (str): Path to Excel file. Uses config default if not provided.
        """
        self.excel_path = Path(excel_path) if excel_path else Config.TEST_DATA_PATH
        self.workbook = None
        self.logger = Logger()

    def load_workbook(self):
        """
        Load the Excel workbook.

        Returns:
            bool: True if loaded successfully, False otherwise
        """
        try:
            if not self.excel_path.exists():
                self.logger.error(f"Excel file not found: {self.excel_path}")
                return False

            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            self.logger.success(f"Excel file loaded: {self.excel_path.name}")
            return True

        except Exception as e:
            self.logger.error(f"Failed to load Excel file: {str(e)}")
            return False

    def read_master_sheet(self):
        """
        Read the Master sheet to get test cases to execute.

        Returns:
            list: List of test case dictionaries or empty list on failure
        """
        try:
            master_sheet_name = Config.MASTER_SHEET_NAME
            if master_sheet_name not in self.workbook.sheetnames:
                self.logger.error(f"Master sheet '{master_sheet_name}' not found in workbook")
                return []

            sheet = self.workbook[master_sheet_name]
            test_cases = []

            # Read header row (row 1)
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(cell.value.strip())

            # Validate required columns
            required_columns = ["TestCaseID", "Execute", "ScreenFlow", "Description"]
            for col in required_columns:
                if col not in headers:
                    self.logger.error(f"Required column '{col}' not found in Master sheet")
                    return []

            # Read test case rows (starting from row 2)
            for row_idx in range(2, sheet.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, start=1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    row_data[header] = cell_value if cell_value is not None else ""

                # Only include test cases marked for execution
                if str(row_data.get("Execute", "")).strip().lower() in ["yes", "y", "true", "1"]:
                    test_cases.append(row_data)

            self.logger.info(f"Found {len(test_cases)} test cases to execute")
            return test_cases

        except Exception as e:
            self.logger.error(f"Error reading Master sheet: {str(e)}")
            return []

    def read_screen_sheet(self, screen_name):
        """
        Read a specific screen sheet.

        Args:
            screen_name (str): Name of the screen sheet

        Returns:
            dict: Screen data with field_names, locators, and test_data
        """
        try:
            if screen_name not in self.workbook.sheetnames:
                self.logger.error(f"Screen sheet '{screen_name}' not found in workbook")
                return None

            sheet = self.workbook[screen_name]

            # Row 1: Field Names
            field_names = []
            for cell in sheet[1]:
                if cell.value:
                    field_names.append(str(cell.value).strip())

            if not field_names:
                self.logger.error(f"No field names found in sheet '{screen_name}'")
                return None

            # Row 2: Locators
            locators = []
            for col_idx in range(1, len(field_names) + 1):
                cell_value = sheet.cell(row=2, column=col_idx).value
                locators.append(str(cell_value).strip() if cell_value else "")

            # Row 3 onwards: Test Data (Keywords and Values in pairs)
            test_data_rows = []
            for row_idx in range(3, sheet.max_row + 1):
                row_data = []
                for col_idx in range(1, len(field_names) + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    row_data.append(str(cell_value).strip() if cell_value is not None else "")

                # Skip empty rows
                if any(row_data):
                    test_data_rows.append(row_data)

            screen_data = {
                "field_names": field_names,
                "locators": locators,
                "test_data": test_data_rows
            }

            self.logger.info(f"Read screen '{screen_name}': {len(field_names)} fields, {len(test_data_rows)} data rows")
            return screen_data

        except Exception as e:
            self.logger.error(f"Error reading screen sheet '{screen_name}': {str(e)}")
            return None

    def get_test_data_for_flow(self, screen_flow):
        """
        Get test data for a screen flow (single or multiple screens).

        Args:
            screen_flow (str): Comma-separated screen names (e.g., "LoginScreen,Dashboard")

        Returns:
            list: List of screen data dictionaries
        """
        try:
            # Split screen flow by comma
            screen_names = [name.strip() for name in screen_flow.split(",") if name.strip()]

            if not screen_names:
                self.logger.error("No screen names provided in screen flow")
                return []

            flow_data = []
            for screen_name in screen_names:
                screen_data = self.read_screen_sheet(screen_name)
                if screen_data:
                    screen_data["screen_name"] = screen_name
                    flow_data.append(screen_data)
                else:
                    self.logger.warning(f"Could not read screen '{screen_name}', skipping")

            return flow_data

        except Exception as e:
            self.logger.error(f"Error getting test data for flow '{screen_flow}': {str(e)}")
            return []

    def parse_test_steps(self, screen_data):
        """
        Parse screen data into executable test steps.

        Args:
            screen_data (dict): Screen data with field_names, locators, and test_data

        Returns:
            list: List of test step dictionaries
        """
        try:
            test_steps = []
            field_names = screen_data.get("field_names", [])
            locators = screen_data.get("locators", [])
            test_data = screen_data.get("test_data", [])

            # Process test data in pairs (keywords row + values row)
            for i in range(0, len(test_data), 2):
                if i + 1 < len(test_data):
                    keywords_row = test_data[i]
                    values_row = test_data[i + 1]

                    # Create steps for each column
                    for col_idx, field_name in enumerate(field_names):
                        keyword = keywords_row[col_idx] if col_idx < len(keywords_row) else ""
                        value = values_row[col_idx] if col_idx < len(values_row) else ""
                        locator = locators[col_idx] if col_idx < len(locators) else ""

                        # Skip if no keyword
                        if not keyword:
                            continue

                        step = {
                            "field_name": field_name,
                            "keyword": keyword.strip().upper(),
                            "locator": locator.strip(),
                            "value": value.strip()
                        }
                        test_steps.append(step)

            return test_steps

        except Exception as e:
            self.logger.error(f"Error parsing test steps: {str(e)}")
            return []

    def close_workbook(self):
        """Close the workbook and release resources."""
        try:
            if self.workbook:
                self.workbook.close()
                self.logger.info("Excel workbook closed")
        except Exception as e:
            self.logger.warning(f"Error closing workbook: {str(e)}")

    def get_all_screen_names(self):
        """
        Get all sheet names from the workbook (excluding Master sheet).

        Returns:
            list: List of screen sheet names
        """
        try:
            if not self.workbook:
                self.logger.error("Workbook not loaded")
                return []

            master_sheet = Config.MASTER_SHEET_NAME
            screen_names = [sheet for sheet in self.workbook.sheetnames if sheet != master_sheet]

            return screen_names

        except Exception as e:
            self.logger.error(f"Error getting screen names: {str(e)}")
            return []

    def validate_excel_structure(self):
        """
        Validate the Excel file structure.

        Returns:
            tuple: (bool, list) - (is_valid, list_of_errors)
        """
        errors = []

        try:
            # Check if Master sheet exists
            master_sheet = Config.MASTER_SHEET_NAME
            if master_sheet not in self.workbook.sheetnames:
                errors.append(f"Master sheet '{master_sheet}' not found")
                return False, errors

            # Validate Master sheet structure
            sheet = self.workbook[master_sheet]
            required_columns = ["TestCaseID", "Execute", "ScreenFlow", "Description"]
            headers = [cell.value for cell in sheet[1] if cell.value]

            for col in required_columns:
                if col not in headers:
                    errors.append(f"Required column '{col}' missing in Master sheet")

            # Validate screen sheets
            screen_names = self.get_all_screen_names()
            for screen_name in screen_names:
                screen_sheet = self.workbook[screen_name]

                # Check if at least 2 rows exist (field names + locators)
                if screen_sheet.max_row < 2:
                    errors.append(f"Screen '{screen_name}' has insufficient rows (needs at least 2)")

            is_valid = len(errors) == 0
            return is_valid, errors

        except Exception as e:
            errors.append(f"Validation error: {str(e)}")
            return False, errors
