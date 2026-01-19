"""
Excel Reader module to parse test data from Excel sheets
"""
import openpyxl
from pathlib import Path
from framework.utils.logger import logger
from framework.config.config import TEST_DATA_CONFIG

class ExcelReader:
    """Reads and parses Excel test data files"""

    def __init__(self, file_path=None):
        """Initialize Excel reader with file path"""
        self.file_path = file_path or TEST_DATA_CONFIG["test_data_file"]
        self.workbook = None
        self.master_sheet_data = []
        self.screen_data = {}

    def load_workbook(self):
        """Load Excel workbook"""
        try:
            if not Path(self.file_path).exists():
                logger.error(f"Excel file not found: {self.file_path}")
                raise FileNotFoundError(f"Excel file not found: {self.file_path}")

            self.workbook = openpyxl.load_workbook(self.file_path)
            logger.info(f"Successfully loaded workbook: {self.file_path}")
            return True
        except Exception as e:
            logger.error(f"Error loading workbook: {str(e)}")
            return False

    def read_master_sheet(self):
        """
        Read Master Sheet to get test case execution configuration
        Expected format:
        Row 1: Headers (TestCaseID, Execute, ScreenFlow, Description)
        Row 2+: Test case data
        """
        try:
            master_sheet_name = TEST_DATA_CONFIG["master_sheet_name"]
            if master_sheet_name not in self.workbook.sheetnames:
                logger.error(f"Master sheet '{master_sheet_name}' not found in workbook")
                return []

            sheet = self.workbook[master_sheet_name]
            logger.info(f"Reading master sheet: {master_sheet_name}")

            # Read headers (first row)
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)

            # Read test case data
            master_data = []
            for row_idx in range(2, sheet.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, start=1):
                    cell_value = sheet.cell(row_idx, col_idx).value
                    row_data[header] = cell_value

                # Only add rows where Execute is 'Yes' or 'Y'
                execute_value = str(row_data.get('Execute', '')).strip().upper()
                if execute_value in ['YES', 'Y']:
                    master_data.append(row_data)
                    logger.info(f"Test case '{row_data.get('TestCaseID')}' selected for execution")

            self.master_sheet_data = master_data
            logger.info(f"Total test cases selected for execution: {len(master_data)}")
            return master_data

        except Exception as e:
            logger.error(f"Error reading master sheet: {str(e)}")
            return []

    def read_screen_sheet(self, screen_name):
        """
        Read screen-specific sheet
        Expected format:
        Row 1: Field names (headers)
        Row 2: XPath/Locators for each field
        Row 3+: Test data (2 rows per test case - keywords and values)
        """
        try:
            if screen_name not in self.workbook.sheetnames:
                logger.error(f"Screen sheet '{screen_name}' not found in workbook")
                return None

            sheet = self.workbook[screen_name]
            logger.info(f"Reading screen sheet: {screen_name}")

            # Read field names (Row 1)
            field_names = []
            for col_idx in range(1, sheet.max_column + 1):
                field_name = sheet.cell(1, col_idx).value
                if field_name:
                    field_names.append(field_name)

            # Read locators (Row 2)
            locators = []
            for col_idx in range(1, len(field_names) + 1):
                locator = sheet.cell(2, col_idx).value
                locators.append(locator)

            # Create field-locator mapping
            field_locator_map = {}
            for idx, field_name in enumerate(field_names):
                field_locator_map[field_name] = locators[idx]

            # Read test case data (Row 3 onwards - 2 rows per test case)
            test_cases = []
            row_idx = 3
            test_case_num = 1

            while row_idx <= sheet.max_row:
                # Read keywords row
                keywords_row = []
                for col_idx in range(1, len(field_names) + 1):
                    keyword = sheet.cell(row_idx, col_idx).value
                    keywords_row.append(keyword if keyword else "")

                # Read values row (next row)
                values_row = []
                if row_idx + 1 <= sheet.max_row:
                    for col_idx in range(1, len(field_names) + 1):
                        value = sheet.cell(row_idx + 1, col_idx).value
                        values_row.append(value if value else "")

                # Create test case data structure
                test_case_data = {
                    "test_case_id": f"{screen_name}_TC{test_case_num}",
                    "steps": []
                }

                for idx, field_name in enumerate(field_names):
                    if keywords_row[idx]:  # Only add if keyword exists
                        step = {
                            "field": field_name,
                            "locator": field_locator_map[field_name],
                            "keyword": keywords_row[idx],
                            "value": values_row[idx] if idx < len(values_row) else ""
                        }
                        test_case_data["steps"].append(step)

                if test_case_data["steps"]:  # Only add if there are steps
                    test_cases.append(test_case_data)

                row_idx += 2  # Move to next test case (skip 2 rows)
                test_case_num += 1

            screen_data = {
                "screen_name": screen_name,
                "field_locator_map": field_locator_map,
                "test_cases": test_cases
            }

            self.screen_data[screen_name] = screen_data
            logger.info(f"Loaded {len(test_cases)} test case(s) from screen '{screen_name}'")
            return screen_data

        except Exception as e:
            logger.error(f"Error reading screen sheet '{screen_name}': {str(e)}")
            return None

    def get_test_data_for_flow(self, screen_flow):
        """
        Get test data for a specific screen flow
        Screen flow is comma-separated list of screens
        Example: "LoginScreen,DashboardScreen,ProfileScreen"
        """
        try:
            screens = [s.strip() for s in screen_flow.split(',')]
            flow_data = []

            for screen in screens:
                if screen not in self.screen_data:
                    screen_data = self.read_screen_sheet(screen)
                    if not screen_data:
                        logger.warning(f"No data found for screen: {screen}")
                        continue
                else:
                    screen_data = self.screen_data[screen]

                flow_data.append(screen_data)

            return flow_data

        except Exception as e:
            logger.error(f"Error getting test data for flow: {str(e)}")
            return []

    def close_workbook(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()
            logger.info("Workbook closed")

    def get_all_test_cases(self):
        """Get all test cases selected for execution"""
        return self.master_sheet_data
